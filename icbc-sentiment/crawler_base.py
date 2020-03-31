# author: yin yalin
# purpose: 构造父类，使用selenium下载相关网站搜索结果
import os
import time
import traceback
from configparser import ConfigParser
from queue import Queue

from openpyxl import load_workbook
from selenium.webdriver import Firefox
from selenium.webdriver.firefox.options import Options

from selenium.webdriver.support.wait import WebDriverWait

import pandas as pd
import download
from logger import CustomLogging, LogType

cp = ConfigParser()
file_path = os.path.dirname(os.path.abspath(__file__))

cp.read(os.path.join(file_path, 'config'))
debug = cp.get('debug', 'visible')


class SentimentCrawler(object):
    def __init__(self, init=True, ):
        if init:
            # option = Options()
            # option.add_argument('--headless')
            # self.driver = Ie()
            # self.driver = Chrome()
            # self.driver = PhantomJS()
            if debug == 1:
                self.driver = Firefox()
            else:
                option = Options()
                option.set_headless(True)
                self.driver = Firefox(options=option)

            self.driver.maximize_window()
            self.wait = WebDriverWait(self.driver, 15)
        self.case_id = '999'
        self.q = Queue()
        self.name = ''
        self.year_range = 1
        self.keyword = ''
        self.titles = []

    def download_and_save_item(self, item):
        '''
        根据item对象里的url下载html页面
        :param item: 新闻文章对应的实体
        :return:
        '''
        print('downloading url: %s' % item.url)
        try:
            if self.name in ('人民网', '中国证券网', '中国经济网', '腾讯财经',) or (
                    self.name == '千龙网' and ('qndj' in item.url or 'bbs' in item.url)):
                ret = download.download_as_html(item.url, encoding='gbk')
                html = ret['text'].encode('gbk', 'ignore')
            elif self.name in ('每经网',):
                ret = download.download_as_html(item.url)
                html = ret['text'].encode('utf-8', 'ignore')
            elif self.name in ('同花顺',):
                ret = download.download_as_html(item.url, encoding='gbk')
                html = ret['text'].encode('gbk', 'ignore')
            else:
                ret = download.download_as_html(item.url)
                if ret['encoding'] == 'ISO-8859-1':
                    html = ret['text'].encode('ISO-8859-1', 'ignore')
                else:
                    html = ret['text']
            if html:
                item.url = ret['url']
                item.full_content = self.parse_html(item.url, html)
                # 存储
                # self.db_handler.save_article_info(item, self.case_id, )
                self.q.put(item)

            else:
                return
        except TypeError:
            CustomLogging.log_to_file(traceback.format_exc(), LogType.INFO)


    def parse_html(self, url, html):
        pass

    def crawl_main_page(self, keyword, ):
        pass

    def crawl_search_results(self):
        pass

    def start_crawl(self, keywords: list or tuple, case_id: str = '999', year_range: int = 1):
        self.year_range = year_range
        self.case_id = case_id
        for keyword in keywords:
            self.keyword = keyword
            self.crawl_main_page(keyword)
            self.driver.quit()

    def write_to_excel(self, filepath):
        # self.db_handler.close_conn()
        article_list = []
        while not self.q.empty():
            item = self.q.get()
            article_list.append(
                (self.name, item.url, item.title, item.short_description, item.publish_date, item.full_content))

        if article_list:
            df = pd.DataFrame(article_list)
            xl_writer = pd.ExcelWriter(filepath, engine='openpyxl')
            if os.path.exists(filepath):
                wb = load_workbook(filepath)
                xl_writer.book = wb
                xl_writer.sheets = dict([(ws.title, ws) for ws in wb.worksheets])
            df.to_excel(xl_writer, index=False, sheet_name=self.name,
                        header=['source', 'url', 'title', 'short_desc', 'issue_date', 'full_content'])
            try:
                xl_writer.save()
            except PermissionError:
                xl_writer_new = pd.ExcelWriter(filepath.replace('.xlsx', '{0}.xlsx'.format(str(int(time.time())))),
                                               engine='openpyxl')
                if os.path.exists(filepath):
                    wb = load_workbook(filepath)
                    xl_writer_new.book = wb
                    xl_writer_new.sheets = dict([(ws.title, ws) for ws in wb.worksheets])
                df.to_excel(xl_writer_new, index=False, sheet_name=self.name,
                            header=['source', 'url', 'title', 'short_desc', 'issue_date', 'full_content'])
                xl_writer_new.save()
        else:
            print('找不到{0}年内任何与关键字 "{1}" 相关的内容'.format(self.year_range, self.keyword))
            CustomLogging.log_to_file('搜索不到匹配记录', LogType.INFO)

        pass


if __name__ == '__main__':
    pass
